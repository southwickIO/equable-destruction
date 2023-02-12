###############################################################################
# NAME: evercookie.py		                                                  #
#                                                                             #
# VERSION: 20230211                                                           #
#                                                                             #
# SYNOPSIS: Main script that runs the other scripts.					      #
#           			                                                      #
#                                                                             #
# DESCRIPTION: This script is part of a larger Evercookie checking            #
#              application for Ubuntu machines. This project was created to   #
#              study evercookies.                                             #
#                                                                             #
# INPUT: None.                                                                #
#                                                                             #
# OUTPUT: 1.) STDOUT                                                          #
#                                                                             #
# PRE-RUNTIME NOTES: 1.) None.				                                  #
#                                                                             #
# AUTHORS: @southwickio                                                       #
#                                                                             #
# LICENSE: GPLv3                                                              #
#                                                                             #
# DISCLAIMER: All work produced by Authors is provided “AS IS”. Authors make  #
#             no warranties, express or implied, and hereby disclaims any and #
#             all warranties, including, but not limited to, any warranty of  #
#             fitness, application, et cetera, for any particular purpose,    #
#             use case, or application of this script.                        #
#                                                                             #
############################################################################### 



#import dependencies
import openpyxl #pip install openpyxl



#stage column headers
headers = ["host_keys", "name", "value", "path", "expires_utc"]



#define functions
def createspreadsheet():
	'''
		spreadsheet init
	'''

	#init workbook/sheet
	workbook = openpyxl.Workbook()
	sheet = workbook.active
	
	#push headers
	for i, header in enumerate(headers):
		sheet.cell(row=1, column=i+1, value=header)

	#save workbook
	workbook.save("../evercookies.xlsx")



def appendspreadsheet(filepath, listofcookies):

	#prep the spreadsheet
	workbook = openpyxl.load_workbook(filepath)
	sheet = workbook.active
	numberofrows = sheet.max_row + 1

	#append the spreadsheet
	for entry in range(len(listofcookies)):
		for i, header in enumerate(headers):
			sheet.cell(row=numberofrows+entry, column=i+1, value=listofcookies[entry][i])

	#save workbook
	workbook.save("../evercookies.xlsx")



#entry point
if __name__ == "__main__":
	print("Please run this application from the main module. Exiting.")
	exit()
else:
	pass
