#!/usr/bin/env python3



###############################################################################
# NAME: spreadsheethandler.py                                                 #
#                                                                             #
# VERSION: 20230221                                                           #
#                                                                             #
# SYNOPSIS: Handles spreadsheet tasks.									      #
#           			                                                      #
#                                                                             #
# DESCRIPTION: This script is part of a larger Evercookie checking            #
#              application for Ubuntu machines. This project was created to   #
#              study evercookies.                                             #
#                                                                             #
# INPUT: None.                                                                #
#                                                                             #
# OUTPUT: 1.) $filepath														  #
#		  2.) stdout			                                              #
#                                                                             #
# PRE-RUNTIME NOTES: 1.) None.				                                  #
#                                                                             #
# AUTHORS: @southwickio                                                       #
#                                                                             #
# LICENSE: GPLv3                                                              #
#                                                                             #
# DISCLAIMER: All work produced by Authors is provided “AS IS”. Authors make  #
#             no warranties, express or implied, and hereby disclaims any and #
#             all warranties, including, but not limited to, any warranty of  #
#             fitness, application, et cetera, for any particular purpose,    #
#             use case, or application of this script.                        #
#                                                                             #
############################################################################### 



#import dependencies
import openpyxl #pip3 install openpyxl



#stage column headers
headers = ["host_keys", "name", "value", "path", "expires_utc"]



#define functions
def createSpreadsheet(filepath):

	#init workbook/sheet
	workbook = openpyxl.Workbook()
	sheet = workbook.active

	#push headers
	for i, header in enumerate(headers):
		sheet.cell(row=1, column=i+1, value=header)

	#save workbook
	workbook.save(filepath)

	return workbook



def stageSpreadsheet(spreadsheetinfo):
	


	return sheet



def addSpace(filepath, times=1):

	#prep the spreadsheet
	workbook = openpyxl.load_workbook(filepath)
	sheet = workbook.active

	#add the space
	for i in range(times):
		sheet.cell(row=sheet.max_row + 1, column=1, value=" ")

	#save workbook
	workbook.save(filepath)



def addTitle(filepath, title):

	#prep the spreadsheet
	workbook = openpyxl.load_workbook(filepath)
	sheet = workbook.active

	#add the title
	sheet.cell(row=sheet.max_row + 1, column=1, value=title)

	#save workbook
	workbook.save(filepath)



def addData(filepath, title, listofcookies):

	#prep the spreadsheet
	workbook = openpyxl.load_workbook(filepath)
	sheet = workbook.active

	numberofrows = sheet.max_row + 1

	#append the data
	for entry in range(len(listofcookies)):
		
		for i, header in enumerate(headers):

			#branch for FindHTMLCookies data
			if "FindHTMLCookies" in title:
				sheet.cell(row=numberofrows + entry, column=i + 1, value = listofcookies[entry][i])
			
			#branch for FindSessionStorage data
			elif "FindSessionStorage" in title:
				pass
				#do stuff

	#save workbook
	workbook.save(filepath)



def appendSpreadsheet(filepath, title, listofcookies):

	addSpace(filepath, 10)
	addTitle(filepath, title)
	addData(filepath, title, listofcookies)



#entry point
if __name__ == "__main__":
	print("Please run this application from the main module. Exiting.")
	exit()
else:
	pass
